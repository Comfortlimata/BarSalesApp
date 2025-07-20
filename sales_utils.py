# sales_utils.py
import sqlite3
from datetime import datetime
from fpdf import FPDF
import hashlib
import os
import json
import bcrypt

DB_NAME = "bar_sales.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    # Create sales table
    cur.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            item TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            price_per_unit REAL NOT NULL,
            total REAL NOT NULL,
            timestamp TEXT NOT NULL
        )
    ''')
    # Create users table
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
    ''')
    # Create inventory table (with category)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            item TEXT PRIMARY KEY,
            quantity INTEGER NOT NULL,
            cost_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            category TEXT DEFAULT ''
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def check_password(password, password_hash):
    return bcrypt.checkpw(password.encode(), password_hash.encode())

def create_user(username, password, role):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    password_hash = hash_password(password)
    try:
        cur.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", (username, password_hash, role))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return False
    conn.close()
    return True

def get_user(username):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT username, password_hash, role FROM users WHERE username=?", (username,))
    row = cur.fetchone()
    conn.close()
    if row:
        return {'username': row[0], 'password_hash': row[1], 'role': row[2]}
    return None

def record_sale(username, item, quantity, price_per_unit):
    from datetime import datetime
    # Stock check
    stock = get_stock(item)
    if quantity > stock:
        raise ValueError(f"Not enough stock for {item}. In stock: {stock}, requested: {quantity}")
    total = quantity * price_per_unit
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_str = datetime.now().strftime("%Y-%m-%d")
    if not timestamp.startswith(now_str):
        raise ValueError("Backdating sales is not allowed.")
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("INSERT INTO sales (username, item, quantity, price_per_unit, total, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
                (username, item, quantity, price_per_unit, total, timestamp))
    conn.commit()
    conn.close()
    # Deduct stock
    update_stock(item, -quantity)
    return total

def export_to_csv():
    import csv
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM sales")
    rows = cur.fetchall()
    conn.close()

    with open("export/sales.csv", "w", newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["ID", "Item", "Quantity", "Unit Price", "Total", "Timestamp"])
        writer.writerows(rows)

def get_total_sales():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT SUM(total) FROM sales")
    total = cur.fetchone()[0]
    conn.close()
    return total if total else 0

def generate_pdf_receipt(receipt_id, username, item, quantity, price_per_unit, total, timestamp):
    business_name = "Comfort_2022 Bar Sales"
    business_address = "123 Main St, Your City"  # Change as needed
    logo_path = os.path.join("assets", "logo.png")
    signature_data = f"{receipt_id}|{username}|{item}|{quantity}|{price_per_unit}|{total}|{timestamp}"
    signature = hashlib.sha256(signature_data.encode()).hexdigest()
    pdf = FPDF()
    pdf.add_page()
    logo_added = False
    if os.path.exists(logo_path):
        try:
            pdf.image(logo_path, x=10, y=8, w=33)
            pdf.set_xy(50, 10)
            logo_added = True
        except Exception:
            # Skip logo if not a valid PNG
            pdf.set_xy(10, 10)
    else:
        pdf.set_xy(10, 10)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, business_name, ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, business_address, ln=1)
    pdf.ln(10)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Receipt ID: {receipt_id}", ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Date/Time: {timestamp}", ln=1)
    pdf.cell(0, 10, f"Cashier: {username}", ln=1)
    pdf.ln(5)
    pdf.cell(0, 10, f"Item: {item}", ln=1)
    pdf.cell(0, 10, f"Quantity: {quantity}", ln=1)
    pdf.cell(0, 10, f"Price per Unit: ZMW {price_per_unit:.2f}", ln=1)
    pdf.cell(0, 10, f"Total: ZMW {total:.2f}", ln=1)
    pdf.ln(10)
    pdf.set_font("Arial", "I", 10)
    pdf.multi_cell(0, 8, f"Digital Signature:\n{signature}")
    if not os.path.exists("exports"):
        os.makedirs("exports")
    pdf_path = os.path.join("exports", f"receipt_{receipt_id}.pdf")
    pdf.output(pdf_path)
    return pdf_path

def backup_today_sales():
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM sales WHERE DATE(timestamp)=?", (today,))
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]
    conn.close()
    sales_list = [dict(zip(columns, row)) for row in rows]
    if not os.path.exists("data"):
        os.makedirs("data")
    backup_path = os.path.join("data", f"backup_{today}.json")
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(sales_list, f, indent=2)
    return backup_path

def log_audit_event(event):
    from datetime import datetime
    if not os.path.exists("data"):
        os.makedirs("data")
    log_path = os.path.join("data", "audit.log")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {event}\n")

def export_all_sales_to_csv(start_date, end_date, selected_columns, export_format='CSV'):
    import csv
    from datetime import datetime
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    col_str = ', '.join(selected_columns)
    query = f"SELECT {col_str} FROM sales WHERE DATE(timestamp) BETWEEN ? AND ?"
    cur.execute(query, (start_date, end_date))
    rows = cur.fetchall()
    conn.close()
    if not os.path.exists("exports"):
        os.makedirs("exports")
    today = datetime.now().strftime("%Y-%m-%d")
    if export_format == 'Excel':
        import openpyxl
        from openpyxl.utils import get_column_letter
        xlsx_path = os.path.join("exports", f"all_sales_{start_date}_to_{end_date}_{today}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(selected_columns)
        for row in rows:
            ws.append(row)
        # Auto-size columns
        for i, col in enumerate(selected_columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)
        wb.save(xlsx_path)
        return xlsx_path
    else:
        csv_path = os.path.join("exports", f"all_sales_{start_date}_to_{end_date}_{today}.csv")
        with open(csv_path, "w", newline='', encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(selected_columns)
            writer.writerows(rows)
        return csv_path

def get_stock(item):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT quantity FROM inventory WHERE item=?', (item,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else 0

def update_stock(item, change):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT quantity FROM inventory WHERE item=?', (item,))
    row = cur.fetchone()
    if row:
        new_qty = row[0] + change
        cur.execute('UPDATE inventory SET quantity=? WHERE item=?', (new_qty, item))
    else:
        cur.execute('INSERT INTO inventory (item, quantity) VALUES (?, ?)', (item, max(0, change)))
    conn.commit()
    conn.close()

def get_all_stock():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT item, quantity, category FROM inventory ORDER BY item')
    rows = cur.fetchall()
    conn.close()
    return rows

def set_item_prices(item, cost, sell):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    if cost is not None and sell is not None:
        cur.execute('UPDATE inventory SET cost_price=?, selling_price=? WHERE item=?', (cost, sell, item))
    elif cost is not None:
        cur.execute('UPDATE inventory SET cost_price=? WHERE item=?', (cost, item))
    elif sell is not None:
        cur.execute('UPDATE inventory SET selling_price=? WHERE item=?', (sell, item))
    conn.commit()
    conn.close()

def set_item_category(item, category):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('UPDATE inventory SET category=? WHERE item=?', (category, item))
    conn.commit()
    conn.close()

def get_item_category(item):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT category FROM inventory WHERE item=?', (item,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else ''

def delete_item(item):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('DELETE FROM inventory WHERE item=?', (item,))
    conn.commit()
    conn.close()

def get_item_prices(item):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT cost_price, selling_price FROM inventory WHERE item=?', (item,))
    row = cur.fetchone()
    conn.close()
    return row if row else (None, None)

def get_sales_history_for_item(item):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute('SELECT timestamp, username, quantity, price_per_unit, total FROM sales WHERE item=? ORDER BY timestamp DESC', (item,))
    rows = cur.fetchall()
    conn.close()
    return rows
